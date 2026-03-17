import pandas as pd
import openpyxl
import numpy as np

# Load the workbook to check formulas
wb = openpyxl.load_workbook('Q_measurement_salt_inj.xlsx')
ws = wb['with_salt']

# Find the discharge cell (should be around row 39, column E)
print("Checking cells around discharge calculation area:")
for row in range(35, 45):
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws[f'{col}{row}']
        if cell.value is not None:
            # Check if it has a formula
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"{col}{row}: FORMULA = {cell.value}")
            else:
                print(f"{col}{row}: {cell.value}")

print("\n" + "="*80)
print("Looking for the Q calculation formula...")
print("="*80)

# Check specifically around the discharge value
for row in range(38, 42):
    for col in ['D', 'E', 'F']:
        cell = ws[f'{col}{row}']
        if cell.value is not None:
            if isinstance(cell.value, str) and '=' in str(cell.value):
                print(f"\n{col}{row} contains formula:")
                print(f"  Value: {cell.value}")
