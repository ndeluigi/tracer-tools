import openpyxl

wb = openpyxl.load_workbook('Q_measurement_salt_inj.xlsx')
ws = wb['with_salt']

print("Cell C27 (calibration coefficient b):")
print(f"Value: {ws['C27'].value}")
print(f"Type: {type(ws['C27'].value)}")

print("\nCell C28 (calibration coefficient a):")
print(f"Value: {ws['C28'].value}")

print("\n" + "="*80)
print("DISCHARGE FORMULA BREAKDOWN:")
print("="*80)
print("""
Formula: Q = M / ((SUM(conductivity) - background * COUNT(conductivity)) * b * dt)

Where:
- M = Mass of salt injected [g] (E35)
- SUM(conductivity) = Sum of all conductivity measurements [µS/cm] (E39:E939)
- background = Background conductivity [µS/cm] (E36)
- COUNT(conductivity) = Number of measurements (E39:E939)
- b = Calibration coefficient [C27] = (g/L) / (µS/cm)
- dt = Time step [s] (E37)

The formula calculates:
1. Total excess conductivity = SUM(conductivity) - background * COUNT(conductivity)
2. Integral of excess concentration = Total excess conductivity * b * dt
3. Discharge Q = M / Integral

Note: The calibration coefficient b is in (g/L)/(µS/cm), not (g/L)/(mS/cm)
This is why we don't need to convert µS/cm to mS/cm!
""")
